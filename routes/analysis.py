# routes/analysis.py - AI-POWERED VERSION with Excel Output

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import json
import sys
import base64
from geopy.distance import geodesic
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sklearn.cluster import DBSCAN
from sklearn.metrics import silhouette_score
from sklearn.linear_model import LinearRegression
import warnings
warnings.filterwarnings('ignore')

def convert_numpy_types(obj):
    """Convert numpy types to Python native types for JSON serialization"""
    if isinstance(obj, np.integer):
        return int(obj)
    elif isinstance(obj, np.floating):
        if np.isnan(obj) or np.isinf(obj):
            return 0
        return float(obj)
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, np.bool_):
        return bool(obj)
    elif isinstance(obj, pd.Timestamp):
        return obj.isoformat()
    elif isinstance(obj, dict):
        return {str(key): convert_numpy_types(value) for key, value in obj.items()}
    elif isinstance(obj, (list, tuple)):
        return [convert_numpy_types(item) for item in obj]
    elif pd.isna(obj):
        return 0
    else:
        return obj

class AIPoweredMiningAnalytics:
    """AI-Powered Mining Analytics with DBSCAN and Linear Regression"""
    
    def __init__(self):
        self.gradient_bands = {
            'Steep Down': (-float('inf'), -8),
            'Mild Down': (-8, -3),
            'Flat': (-3, 3),
            'Mild Up': (3, 8),
            'Steep Up': (8, float('inf'))
        }
        self.diesel_price = 94.5
        self.start_radius = 50  # meters for trip detection
        self.eps = 0.0003  # DBSCAN sensitivity
        self.gradient_limit = 10  # pitch threshold
        
    def calculate_gps_distance(self, df):
        """Calculate GPS distance between consecutive points"""
        distances = [0]
        
        for i in range(1, len(df)):
            try:
                p1 = (float(df.iloc[i-1]['lat']), float(df.iloc[i-1]['lon']))
                p2 = (float(df.iloc[i]['lat']), float(df.iloc[i]['lon']))
                
                dist = geodesic(p1, p2).meters
                if dist < 1:
                    distances.append(0)
                else:
                    distances.append(dist)
            except:
                distances.append(0)
                
        df['distance_m'] = distances
        df['distance_km'] = df['distance_m'] / 1000
        return df
    
    def classify_gradient(self, pitch):
        """Classify gradient based on pitch angle"""
        if pd.isna(pitch):
            return 'Flat'
        try:
            for band_name, (low, high) in self.gradient_bands.items():
                if low < pitch <= high:
                    return band_name
        except:
            pass
        return 'Flat'
    
    def detect_start_points_with_dbscan(self, df):
        """Use DBSCAN to automatically detect start/end points"""
        coords = df[['lat', 'lon']].values
        
        # Apply DBSCAN clustering
        db = DBSCAN(eps=self.eps, min_samples=20).fit(coords)
        df['cluster'] = db.labels_
        
        # Filter out noise (-1)
        valid = df[df['cluster'] != -1]
        
        if len(valid) > 0 and len(valid['cluster'].unique()) > 1:
            try:
                score = silhouette_score(coords[df['cluster'] != -1], 
                                        df[df['cluster'] != -1]['cluster'])
                print(f"  Clustering Quality: {round(score, 3)}", file=sys.stderr)
            except:
                pass
        
        # Find main start cluster (most frequent)
        if len(valid) > 0:
            start_cluster = valid['cluster'].value_counts().idxmax()
            start_points = df[df['cluster'] == start_cluster]
            
            start_lat = start_points['lat'].mean()
            start_lon = start_points['lon'].mean()
            
            return (start_lat, start_lon)
        else:
            # Fallback: use first point
            return (df.iloc[0]['lat'], df.iloc[0]['lon'])
    
    def detect_trips_with_ai(self, df):
        """AI-powered trip detection using DBSCAN and radius"""
        # First, detect start point using DBSCAN
        start_point = self.detect_start_points_with_dbscan(df)
        print(f"  AI Detected Start Point: {start_point}", file=sys.stderr)
        
        # Now detect trips based on distance from start point
        trips = []
        current_trip = 0
        in_trip = False
        
        for i in range(len(df)):
            cur_point = (df.iloc[i]['lat'], df.iloc[i]['lon'])
            dist_from_start = geodesic(cur_point, start_point).meters
            
            # Start a trip when we leave the start area
            if dist_from_start > self.start_radius and not in_trip:
                current_trip += 1
                in_trip = True
                trips.append(current_trip)
            # End trip when we return to start area
            elif dist_from_start < self.start_radius and in_trip and i > 50:
                in_trip = False
                trips.append(current_trip)
            else:
                trips.append(current_trip if in_trip else 0)
        
        df['trip'] = trips
        df['start_point_lat'] = start_point[0]
        df['start_point_lon'] = start_point[1]
        
        return df
    
    def train_fuel_prediction_model(self, df):
        """Train Linear Regression model to predict fuel consumption"""
        # Prepare features
        features = ['distance_m', 'pitch', 'alt', 'rl']
        X = df[features].fillna(0)
        y = df['fuel'].fillna(0)
        
        # Train model
        model = LinearRegression()
        model.fit(X, y)
        
        # Calculate R² score
        r2_score = model.score(X, y)
        print(f"  Fuel Model R²: {round(r2_score, 3)}", file=sys.stderr)
        
        # Get feature importance
        importance = dict(zip(features, model.coef_))
        
        return model, importance, r2_score
    
    def analyze_gradient_sections(self, df):
        """Analyze gradient vs flat sections"""
        gradient = df[df['pitch'].abs() > self.gradient_limit]
        flat = df[df['pitch'].abs() <= self.gradient_limit]
        
        grad_stats = {
            'points': len(gradient),
            'distance_km': round(gradient['distance_km'].sum(), 2),
            'fuel_l': round(gradient['fuel'].sum(), 2),
            'cost_rs': round(gradient['cost'].sum(), 2) if 'cost' in gradient.columns else 0,
            'avg_pitch': round(gradient['pitch'].mean(), 1) if len(gradient) > 0 else 0
        }
        
        flat_stats = {
            'points': len(flat),
            'distance_km': round(flat['distance_km'].sum(), 2),
            'fuel_l': round(flat['fuel'].sum(), 2),
            'cost_rs': round(flat['cost'].sum(), 2) if 'cost' in flat.columns else 0,
            'avg_pitch': round(flat['pitch'].mean(), 1) if len(flat) > 0 else 0
        }
        
        return grad_stats, flat_stats
    
    def analyze_device(self, device_id, df):
        """Complete AI-powered analysis for a single device"""
        if len(df) < 10:
            return None
        
        try:
            # Calculate GPS distances
            df = self.calculate_gps_distance(df)
            
            # Classify gradients
            df['gradient_class'] = df['pitch'].apply(self.classify_gradient)
            
            # AI-powered trip detection
            df = self.detect_trips_with_ai(df)
            
            # Train fuel prediction model
            model, feature_importance, r2_score = self.train_fuel_prediction_model(df)
            
            # Calculate AI predictions
            features = ['distance_m', 'pitch', 'alt', 'rl']
            X_pred = df[features].fillna(0)
            df['predicted_fuel'] = model.predict(X_pred)
            
            # Basic metrics
            total_distance = float(df['distance_km'].sum())
            total_fuel = float(df['fuel'].sum())
            total_predicted_fuel = float(df['predicted_fuel'].sum())
            
            # Trip analysis
            unique_trips = df[df['trip'] > 0]['trip'].nunique()
            trip_details = []
            
            for trip_num in sorted(df[df['trip'] > 0]['trip'].unique()):
                trip_data = df[df['trip'] == trip_num]
                if len(trip_data) < 5:
                    continue
                    
                grad_stats, flat_stats = self.analyze_gradient_sections(trip_data)
                
                trip_details.append({
                    'trip_number': int(trip_num),
                    'start_time': str(trip_data.iloc[0]['time']),
                    'end_time': str(trip_data.iloc[-1]['time']),
                    'duration': str(trip_data.iloc[-1]['time'] - trip_data.iloc[0]['time']),
                    'distance_km': round(float(trip_data['distance_km'].sum()), 2),
                    'fuel_l': round(float(trip_data['fuel'].sum()), 2),
                    'predicted_fuel_l': round(float(trip_data['predicted_fuel'].sum()), 2),
                    'gradient_distance_km': grad_stats['distance_km'],
                    'flat_distance_km': flat_stats['distance_km'],
                    'avg_speed': round(float(trip_data['speed'].mean()), 1) if 'speed' in trip_data.columns else 0
                })
            
            # Average speed
            avg_speed = 0
            if 'speed' in df.columns:
                moving = df[df['speed'] > 2]
                avg_speed = float(moving['speed'].mean()) if len(moving) > 0 else 0
            
            # Fuel efficiency
            fuel_per_km = total_fuel / total_distance if total_distance > 0 else 0
            
            # Gradient analysis
            gradient_stats = {}
            for gradient in self.gradient_bands.keys():
                mask = df['gradient_class'] == gradient
                grad_data = df[mask]
                
                if len(grad_data) > 0:
                    distance = float(grad_data['distance_km'].sum())
                    fuel = float(grad_data['fuel'].sum())
                    
                    gradient_stats[gradient] = {
                        'distance_km': round(distance, 2),
                        'fuel_l': round(fuel, 2),
                        'fuel_per_km': round(fuel/distance if distance > 0 else 0, 3),
                        'percentage': round(distance/total_distance * 100, 1) if total_distance > 0 else 0
                    }
                else:
                    gradient_stats[gradient] = {
                        'distance_km': 0,
                        'fuel_l': 0,
                        'fuel_per_km': 0,
                        'percentage': 0
                    }
            
            # AI Metrics
            ai_metrics = {
                'model_r2_score': round(r2_score, 3),
                'feature_importance': {k: round(v, 4) for k, v in feature_importance.items()},
                'total_predicted_fuel': round(total_predicted_fuel, 2),
                'prediction_accuracy': round((1 - abs(total_fuel - total_predicted_fuel) / total_fuel) * 100, 1) if total_fuel > 0 else 0,
                'detected_start_point': f"({df.iloc[0]['start_point_lat']:.6f}, {df.iloc[0]['start_point_lon']:.6f})"
            }
            
            # Idle analysis
            idle_points = len(df[df['speed'] <= 2]) if 'speed' in df.columns else 0
            idle_ratio = idle_points / len(df) if len(df) > 0 else 0
            
            if idle_ratio > 0.4:
                idle_pattern = "High"
            elif idle_ratio > 0.2:
                idle_pattern = "Medium"
            else:
                idle_pattern = "Low"
            
            # Route type
            steep_up_pct = gradient_stats.get('Steep Up', {}).get('percentage', 0)
            
            if steep_up_pct > 30:
                route_type = "Route B (Steep)"
                route_chars = "Steep ramps, high fuel consumption"
            elif steep_up_pct > 15:
                route_type = "Mixed"
                route_chars = "Balanced terrain, moderate efficiency"
            else:
                route_type = "Route A (Gentle)"
                route_chars = "Gentle slopes, good efficiency"
            
            # Savings potential
            target_fuel_per_km = 0.55
            fuel_diff = max(0, fuel_per_km - target_fuel_per_km)
            
            return {
                'device_id': str(device_id),
                'total_distance': round(total_distance, 1),
                'total_fuel': round(total_fuel, 1),
                'fuel_per_km': round(fuel_per_km, 2),
                'trips': unique_trips,
                'dumps': unique_trips,
                'avg_speed': round(avg_speed, 1),
                'speed_interpretation': 'Ramp-limited climbs' if avg_speed < 3 else 'Efficient cycles',
                'gradient_stats': gradient_stats,
                'idle_pattern': idle_pattern,
                'route_type': route_type,
                'route_chars': route_chars,
                'steep_up_percentage': round(steep_up_pct, 1),
                'fuel_diff_per_km': round(fuel_diff, 2),
                'ai_metrics': ai_metrics,
                'trip_details': trip_details[:10]  # Top 10 trips
            }
            
        except Exception as e:
            print(f"Error analyzing device {device_id}: {str(e)}", file=sys.stderr)
            return None

class ExcelReportGenerator:
    """Generate Excel report with AI analytics"""
    
    def __init__(self):
        self.wb = Workbook()
        self.setup_styles()
        
    def setup_styles(self):
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.header_font = Font(color='FFFFFF', bold=True)
        self.centered = Alignment(horizontal='center', vertical='center')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def create_per_hauler_summary(self, results):
        ws = self.wb.create_sheet("Per_Hauler_Summary", 0)
        
        headers = ['Hauler', 'Distance (km)', 'Fuel (L)', 'Fuel/km', 'Trips', 'Avg Speed']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.centered
            cell.border = self.border
        
        row = 2
        for device_id, data in results.items():
            ws.cell(row=row, column=1, value=device_id)
            ws.cell(row=row, column=2, value=data['total_distance'])
            ws.cell(row=row, column=3, value=data['total_fuel'])
            ws.cell(row=row, column=4, value=data['fuel_per_km'])
            ws.cell(row=row, column=5, value=data['trips'])
            ws.cell(row=row, column=6, value=data['avg_speed'])
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = self.border
            row += 1
        
        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        return ws
    
    def create_ai_insights_sheet(self, results):
        """New sheet for AI-powered insights"""
        ws = self.wb.create_sheet("AI_Insights")
        
        headers = ['Hauler', 'Model R²', 'Prediction Accuracy %', 'Detected Start Point', 
                   'Feature: distance', 'Feature: pitch', 'Feature: alt', 'Feature: rl']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.centered
            cell.border = self.border
        
        row = 2
        for device_id, data in results.items():
            ai = data.get('ai_metrics', {})
            importance = ai.get('feature_importance', {})
            
            ws.cell(row=row, column=1, value=device_id)
            ws.cell(row=row, column=2, value=ai.get('model_r2_score', 0))
            ws.cell(row=row, column=3, value=ai.get('prediction_accuracy', 0))
            ws.cell(row=row, column=4, value=ai.get('detected_start_point', 'N/A'))
            ws.cell(row=row, column=5, value=importance.get('distance_m', 0))
            ws.cell(row=row, column=6, value=importance.get('pitch', 0))
            ws.cell(row=row, column=7, value=importance.get('alt', 0))
            ws.cell(row=row, column=8, value=importance.get('rl', 0))
            
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = self.border
            row += 1
        
        for col in range(1, 9):
            ws.column_dimensions[chr(64 + col)].width = 18
        
        return ws
    
    def create_trip_analysis_sheet(self, results):
        """New sheet for detailed trip analysis"""
        ws = self.wb.create_sheet("Trip_Analysis")
        
        headers = ['Hauler', 'Trip #', 'Distance (km)', 'Fuel (L)', 'Predicted Fuel (L)', 
                   'Gradient km', 'Flat km', 'Avg Speed', 'Duration']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.centered
            cell.border = self.border
        
        row = 2
        for device_id, data in results.items():
            for trip in data.get('trip_details', []):
                ws.cell(row=row, column=1, value=device_id)
                ws.cell(row=row, column=2, value=trip.get('trip_number', 0))
                ws.cell(row=row, column=3, value=trip.get('distance_km', 0))
                ws.cell(row=row, column=4, value=trip.get('fuel_l', 0))
                ws.cell(row=row, column=5, value=trip.get('predicted_fuel_l', 0))
                ws.cell(row=row, column=6, value=trip.get('gradient_distance_km', 0))
                ws.cell(row=row, column=7, value=trip.get('flat_distance_km', 0))
                ws.cell(row=row, column=8, value=trip.get('avg_speed', 0))
                ws.cell(row=row, column=9, value=trip.get('duration', '0'))
                
                for col in range(1, 10):
                    ws.cell(row=row, column=col).border = self.border
                row += 1
        
        for col in range(1, 10):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        return ws
    
    def create_gradient_sheet(self, device_id, data):
        ws = self.wb.create_sheet(f"Gradient_{device_id}")
        
        headers = ['Gradient', 'Distance (km)', 'Fuel (L)', 'Fuel/km', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.centered
            cell.border = self.border
        
        row = 2
        gradient_order = ['Steep Down', 'Mild Down', 'Flat', 'Mild Up', 'Steep Up']
        
        for gradient in gradient_order:
            stats = data['gradient_stats'].get(gradient, {})
            ws.cell(row=row, column=1, value=gradient)
            ws.cell(row=row, column=2, value=stats.get('distance_km', 0))
            ws.cell(row=row, column=3, value=stats.get('fuel_l', 0))
            ws.cell(row=row, column=4, value=stats.get('fuel_per_km', 0))
            ws.cell(row=row, column=5, value=stats.get('percentage', 0))
            
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.border
            row += 1
        
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        return ws
    
    def create_speed_analysis(self, results):
        ws = self.wb.create_sheet("Speed_Analysis")
        
        headers = ['Hauler', 'Avg Speed', 'Interpretation']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.centered
            cell.border = self.border
        
        row = 2
        for device_id, data in results.items():
            ws.cell(row=row, column=1, value=device_id)
            ws.cell(row=row, column=2, value=data.get('avg_speed', 0))
            ws.cell(row=row, column=3, value=data.get('speed_interpretation', 'Unknown'))
            
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.border
            row += 1
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        
        return ws
    
    def create_route_comparison(self):
        ws = self.wb.create_sheet("Route_Comparison")
        
        data = [
            ['Metric', 'Route A (Gentle)', 'Route B (Steep)'],
            ['Distance', 'Longer', 'Shorter'],
            ['Max Gradient', 'Low–moderate', 'High (10–27°)'],
            ['Fuel/km', '0.50–0.55', '0.85–0.95'],
            ['Fuel/trip', 'Lower', 'Higher'],
            ['Avg Speed', '5–6 km/h', '1.5–2 km/h'],
            ['Trips/day', 'Higher', 'Lower']
        ]
        
        for r_idx, row_data in enumerate(data, 1):
            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.centered
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        
        return ws
    
    def create_route_usage(self, results):
        ws = self.wb.create_sheet("Route_Usage")
        
        headers = ['Hauler', 'Primary Route', 'Characteristics']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.centered
            cell.border = self.border
        
        row = 2
        for device_id, data in results.items():
            ws.cell(row=row, column=1, value=device_id)
            ws.cell(row=row, column=2, value=data.get('route_type', 'Unknown'))
            ws.cell(row=row, column=3, value=data.get('route_chars', 'Unknown'))
            
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.border
            row += 1
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        
        return ws
    
    def create_idle_analysis(self, results):
        ws = self.wb.create_sheet("Idle_Analysis")
        
        headers = ['Hauler', 'Idle Pattern']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.centered
            cell.border = self.border
        
        row = 2
        for device_id, data in results.items():
            ws.cell(row=row, column=1, value=device_id)
            ws.cell(row=row, column=2, value=data.get('idle_pattern', 'Low'))
            
            for col in range(1, 3):
                ws.cell(row=row, column=col).border = self.border
            row += 1
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        
        return ws
    
    def create_fuel_cost(self, results):
        ws = self.wb.create_sheet("Fuel_Cost")
        
        headers = ['Hauler', 'Fuel (L)', 'Price/L', 'Cost (₹)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.centered
            cell.border = self.border
        
        row = 2
        diesel_price = 94.5
        total_fuel = 0
        total_cost = 0
        
        for device_id, data in results.items():
            fuel = data.get('total_fuel', 0)
            cost = fuel * diesel_price
            
            ws.cell(row=row, column=1, value=device_id)
            ws.cell(row=row, column=2, value=round(fuel, 1))
            ws.cell(row=row, column=3, value=diesel_price)
            ws.cell(row=row, column=4, value=round(cost))
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = self.border
            
            total_fuel += fuel
            total_cost += cost
            row += 1
        
        # Total row
        ws.cell(row=row, column=1, value="TOTAL")
        ws.cell(row=row, column=2, value=round(total_fuel, 1))
        ws.cell(row=row, column=3, value="")
        ws.cell(row=row, column=4, value=round(total_cost))
        
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = self.border
            if col == 1:
                cell.font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        return ws
    
    def generate_report(self, results):
        """Generate complete Excel report with all sheets"""
        
        if 'Sheet' in self.wb.sheetnames:
            std = self.wb['Sheet']
            self.wb.remove(std)
        
        self.create_per_hauler_summary(results)
        self.create_ai_insights_sheet(results)
        self.create_trip_analysis_sheet(results)
        
        for device_id, data in results.items():
            self.create_gradient_sheet(device_id, data)
        
        self.create_speed_analysis(results)
        self.create_route_comparison()
        self.create_route_usage(results)
        self.create_idle_analysis(results)
        self.create_fuel_cost(results)
        
        return self.wb

def main():
    """Main entry point"""
    try:
        input_data = sys.stdin.read()
        if not input_data:
            print(json.dumps({'status': 'error', 'error': 'No input data'}))
            return
        
        params = json.loads(input_data)
        df = pd.DataFrame(params.get('data', []))
        
        if df.empty:
            print(json.dumps({
                'status': 'success',
                'report': '',
                'filename': 'No_Data.xlsx',
                'devices_discovered': [],
                'record_count': 0
            }))
            return
        
        # Ensure required columns
        required = ['time', 'lat', 'lon', 'pitch', 'fuel', 'speed', 'alt']
        for col in required:
            if col not in df.columns:
                df[col] = 0
        
        # Add RL if not present
        if 'rl' not in df.columns:
            df['rl'] = df['alt'] + 525.5  # Sea level constant
        
        # Convert time
        df['time'] = pd.to_datetime(df['time'])
        df = df.sort_values('time').reset_index(drop=True)
        
        print(f"Processing {len(df)} records", file=sys.stderr)
        
        # Initialize AI analyzer
        analyzer = AIPoweredMiningAnalytics()
        
        # Analyze each device
        devices = df['device_id'].unique() if 'device_id' in df.columns else ['unknown']
        results = {}
        
        for device_id in devices:
            if pd.isna(device_id):
                continue
            
            device_df = df[df['device_id'] == device_id].copy()
            
            if len(device_df) < 10:
                print(f"Device {device_id}: insufficient data", file=sys.stderr)
                continue
            
            result = analyzer.analyze_device(device_id, device_df)
            if result:
                results[str(device_id)] = result
                print(f"Device {device_id}: {result['trips']} trips, R²={result['ai_metrics']['model_r2_score']}", file=sys.stderr)
        
        if not results:
            print(json.dumps({
                'status': 'success',
                'report': '',
                'filename': 'No_Valid_Data.xlsx',
                'devices_discovered': [],
                'record_count': len(df)
            }))
            return
        
        # Generate Excel
        generator = ExcelReportGenerator()
        wb = generator.generate_report(results)
        
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        # Create filename
        date_str = datetime.now().strftime("%d%b%y").upper()
        filename = f"AI_Mining_Analytics_{date_str}.xlsx"
        
        # Output
        output = {
            'report': base64.b64encode(excel_bytes.read()).decode('utf-8'),
            'filename': filename,
            'status': 'success',
            'devices_discovered': list(results.keys()),
            'record_count': len(df)
        }
        
        print(json.dumps(output))
        
    except Exception as e:
        import traceback
        print(json.dumps({
            'status': 'error',
            'error': str(e),
            'traceback': traceback.format_exc()
        }), file=sys.stderr)

if __name__ == '__main__':
    main()